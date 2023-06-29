import numpy as np
from itertools import chain, combinations
import time
import pandas as pd
import sys
import matplotlib.pyplot as plt
sys.setrecursionlimit(100000)
import imageio
import matplotlib.ticker as ticker
import random
from openpyxl import Workbook, load_workbook

# ------------------------------- Main Functions ---------------------------------
def weighted_increasing_subsequence_iter(input):
    n = len(input)
    memo = [0] * n
    results = {}

    for i in range(n - 1, -1, -1):
        x, y, w = input[i]
        memo[i] = w
        results[input[i]] = w
        for j in range(i + 1, n):
            x1, y1, w1 = input[j]
            if x < x1 and y < y1:
                if memo[i] < memo[j] + w:
                    memo[i] = memo[j] + w
                    results[input[i]] = memo[i]

    return max(memo), results


def weighted_increasing_subsequence_recursion(input, index=0, largestA=None, memo={}, results = {}):
    if index >= len(input):
        return 0, results
    if (index, largestA) in memo:
        return memo[(index, largestA)], results
    x, y, w = input[index]
    if largestA is not None and (x <= largestA[0] or y <= largestA[1]):
        result, _ = weighted_increasing_subsequence_recursion(input, index + 1, largestA, memo)
    else:
        without_me, _ = weighted_increasing_subsequence_recursion(input, index + 1, largestA, memo)
        with_me, _ = weighted_increasing_subsequence_recursion(input, index + 1, (x, y), memo)
        with_me = with_me + w
        results[input[index]] = with_me
        result = max(without_me, with_me)
    memo[(index, largestA)] = result
    return result, results


def naive_approach(input_list, original_wis):
    input_array = np.array(input_list)
    subsets = list(chain(*map(lambda x: combinations(range(len(input_list)), x), range(0, len(input_list)+1))))
    max_changed = 0
    i = 0
    for subset in subsets:
        subset_array = input_array.copy()
        for index in subset:
            subset_array[index, 2] = 2
        sum_of_third_elements = np.sum(subset_array[:, 2])
        tmp_changed = sum_of_third_elements - len(input_list)
        subset_array = list(map(tuple, subset_array))
        memo = {}
        results = {}
        tmp_wis , _ = weighted_increasing_subsequence_recursion(subset_array)
        memo.clear()
        results.clear()
        if tmp_changed == 5 and tmp_wis == 5:
            i = i +1
        if tmp_wis == original_wis:
            if max_changed < tmp_changed:
                max_changed = tmp_changed
    return max_changed


def heuristic_algo(input_list, ranked_input, maximum_wis, y_dictionary):
    tmp_input_list = list(input_list)
    modified_list = list(input_list)
    changed_items = []
    addition = 0
    i = 0

    while len(tmp_input_list) != 0:
        lowest_partial_order, tmp_list = find_lowest_partial_order(tmp_input_list)
        tmp_input_list = tmp_list
        flag = False

        for elem in lowest_partial_order:
            if ranked_input[elem] + addition < maximum_wis:
                (x, y, w) = elem
                changed_items.append((x, y_dictionary.get(x), w + 1))
                modified_list[modified_list.index(elem)] = (x, y, w + 1)
                flag = True
                # plot_points(input_list, changed_items, i)  # Here we call the plot function.
                i += 1

        addition = addition + 2 if flag else addition + 1
    return changed_items, modified_list


# ------------------------------- Helper Functions ---------------------------------
def create_or_overwrite_excel(sheetname, data):
    # Create a new workbook or load an existing one
    try:
        wb = load_workbook('changed_data.xlsx')
    except FileNotFoundError:
        wb = Workbook()

    # Check if the sheet already exists
    if sheetname in wb.sheetnames:
        # Remove the existing sheet
        wb.remove(wb[sheetname])

    # Create a new sheet
    sheet = wb.create_sheet(title=sheetname, index=0)

    # Add column headers
    sheet["A1"] = "Column 1"
    sheet["B1"] = "Column 2"

    # Add data rows
    for i, (x, y, w) in enumerate(data, start=2):
        sheet[f"A{i}"] = x
        sheet[f"B{i}"] = y

    # Save the workbook
    wb.save('changed_data.xlsx')

def execute(input_list, sheet_name):
    y_dictionary =  {x: y for x, y, _ in input_list}
    input_list = reduction_to_third_dimension(input_list)
    maximum_wis, ranked_input = weighted_increasing_subsequence_iter(input_list)
    print("Maximum WIS in original input: {}".format(maximum_wis))
    # print("Ranked Input: " + str(ranked_input))
    arbiv_start_time = time.time()
    changed_items, modified_list = heuristic_algo(input_list, ranked_input, maximum_wis, y_dictionary)
    arbiv_end_time = time.time()

    create_or_overwrite_excel(sheet_name, changed_items)

    # Finally, compile your saved plots into a gif using imageio.
    # images = []
    # filenames = [f'plot_{i}.png' for i in range(len(changed_items))]
    #
    # for filename in filenames:
    #     images.append(imageio.imread(filename))
    # imageio.mimsave('output.gif', images, duration=0.1, loop=0)

    # print("The Algorithm changed the following items: {}".format(changed_items))
    # print("Modified List: {}".format(modified_list))
    new_maximum_wis, ranked_input = weighted_increasing_subsequence_iter(modified_list)
    # print(ranked_input)
    if new_maximum_wis == maximum_wis:
        print("The Algorithm Passed Within {:.3f} Seconds! \n"
              "The Algorithm succeeded in changing {} out of {} items".format(arbiv_end_time - arbiv_start_time,
                                                                              len(changed_items), len(input_list)))

        return True
    else:
        print("The Algorithm Failed!\n"
              "New Maximum WIS: {} , Original Maximum WIS: {}".format(new_maximum_wis, maximum_wis))
        return False
    # naive_start_time = time.time()
    # print("Naive Approach Changed {} Items".format(naive_approach(input_list, maximum_wis)))
    # native_end_time = time.time()
    # print("Naive Approach O(N^2 * 2^N) Took {} Second!".format(native_end_time - naive_start_time))


def find_lowest_partial_order(input_list):
    lowest_partial_order = []
    to_remove = []
    for (x, y, w) in input_list:
        flag = True
        for (x_low, y_low, _) in lowest_partial_order:
            if x_low < x and y_low < y:
                flag = False
                break
        if flag:
            lowest_partial_order.append((x, y, w))
            to_remove.append((x, y, w))
    for item in to_remove:
        input_list.remove(item)
    return lowest_partial_order, input_list


def read_xlsx_file(sheet_name, num_of_points):
    df = pd.read_excel('data.xlsx', sheet_name=sheet_name)
    points =  [(df['Column 1'][x], df['Column 2'][x], 1) for x in range(num_of_points)]

    x_values = [point[0] for point in points]
    y_values = [point[1] for point in points]

    # plot_original_points(sheet_name, x_values, y_values)

    return points


# ------------------------------- Functions for plotting and gifs -------------------------------
def plot_original_points(sheet_name, x_values, y_values):
    plt.scatter(x_values, y_values)
    plt.xlabel('Column 1')
    plt.ylabel('Column 2')
    plt.title('Original: ' + sheet_name)
    plt.show()

def plot_points(points, changed_items, i):
    # Clear the previous plot.
    plt.clf()

    # Plot all points in blue.
    for point in points:
        plt.plot(point[0], point[1], 'bo')

    # Overplot the changed points in red.
    for changed_point in changed_items:
        plt.plot(changed_point[0], changed_point[1], 'ro')

    # Save each plot with a different name (sequentially).
    plt.savefig(f'plot_{i}.png')


def remove_third_entry(input_list):
    return [(x, y) for x, y, _ in input_list]


def plot_chain(input_points, chain):
    input_points = remove_third_entry(input_points)

    # Split input points and chain into X and Y coordinates for plotting
    input_x, input_y = zip(*input_points)
    chain_x, chain_y, _ = zip(*chain)

    # Calculate weight of the chain
    weight = sum(w for _, _, w in chain)

    # Create the plot
    plt.scatter(input_x, input_y, color='blue')
    plt.plot(chain_x, chain_y, color='red', linewidth=2)
    plt.title('Points and Heaviest Chain')
    plt.xlabel('X')
    plt.ylabel('Y')

    # Annotate weight of the heaviest chain
    mid_point_index = len(chain) // 2
    plt.annotate(f'Weight: {weight}', xy=(chain_x[mid_point_index], chain_y[mid_point_index]), xytext=(0, 10),
                 textcoords='offset points', ha='center', va='bottom',
                 bbox=dict(boxstyle='round,pad=0.2', fc='yellow', alpha=0.3),
                 arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0.5', color='black'))

    plt.show()


def weighted_increasing_subsequence_plot_version(input_list):
    n = len(input_list)
    memo = [0] * n
    results = {}
    chains = {}

    for i in range(n - 1, -1, -1):
        x, y, w = input_list[i]
        memo[i] = w
        results[input_list[i]] = w
        chains[input_list[i]] = [input_list[i]]
        for j in range(i + 1, n):
            x1, y1, w1 = input_list[j]
            if x < x1 and y < y1:
                if memo[i] < memo[j] + w:
                    memo[i] = memo[j] + w
                    results[input_list[i]] = memo[i]
                    chains[input_list[i]] = [input_list[i]] + chains[input_list[j]]

    # Find the starting point of the maximum weighted chain
    max_weight_point = max(results, key=results.get)

    return max(memo), results, chains[max_weight_point]


def visualize_weighted_increasing_algorithm(input_list, maximum_wis, ranked_input, max_chain):
    images = []
    chain_images = []
    maximum_wis_list = [i for i in range(maximum_wis, 1, -1)]

    fig, ax = plt.subplots()
    colors = np.arange(len(input_list))

    ind = 0

    # Scatter plot for the points
    # Traverse the ranked_input dictionary
    for i, ((x, y, w), max_chain_weight) in enumerate(ranked_input.items()):

        # Scatter plot for the points
        scatter = ax.scatter([p[0] for p in input_list], [p[1] for p in input_list],
                             c=colors, cmap='viridis', s=[p[2] * 100 for p in input_list], alpha=0.6)

        # Highlight the current point
        ax.scatter([x], [y], color='red', s=w * 150)

        # Annotate the points with their weights and max_chain_weights
        for (x_, y_, w_), max_chain_weight_ in ranked_input.items():
            ax.text(x_, y_, f'w:{w_}, mcw:{max_chain_weight_}', fontsize=8)

        # Set titles, labels and legend
        ax.set_title(f'Step {i + 1}: Point ({x},{y}) with weight {w} and max chain weight {max_chain_weight}\n')
        ax.set_xlabel('X')
        ax.set_ylabel('Y')
        legend1 = ax.legend(*scatter.legend_elements(num=len(input_list)),
                            loc="upper left", title="Points")
        ax.add_artist(legend1)

        # Save the figure as an image file
        fig.canvas.draw()  # draw the canvas, cache the renderer
        image = np.frombuffer(fig.canvas.tostring_rgb(), dtype='uint8')
        image = image.reshape(fig.canvas.get_width_height()[::-1] + (3,))
        images.append(image)

        # If the point is part of the maximum chain, plot the chain up to this point
        if (x, y, w) in max_chain:
            chain_index = max_chain.index((x, y, w))
            if chain_index > 0:
                ax.set_title(f'Current Maximum Chain Weight: {maximum_wis_list[ind]}')
                ind += 1
                chain_x = [p[0] for p in max_chain[:chain_index + 1]]
                chain_y = [p[1] for p in max_chain[:chain_index + 1]]
                ax.plot(chain_x, chain_y, 'b-')
                fig.canvas.draw()
                image = np.frombuffer(fig.canvas.tostring_rgb(), dtype='uint8')
                image = image.reshape(fig.canvas.get_width_height()[::-1] + (3,))
                chain_images.append(image)

        # Remove texts and points for next iteration
        ax.clear()

    imageio.mimsave('algorithm_visualization.gif', images + chain_images[::-1], duration=1500, loop=0)
    print("GIF has been saved as 'algorithm_visualization.gif'")


def find_heaviest_chain(points):
    # Initialize the DP table with the weights of the points
    dp = [w for x, y, w in points]

    # Iterate over all points
    for i in range(1, len(points)):
        for j in range(i):
            # Check if the i-th point is greater than the j-th point
            if points[i][0] > points[j][0] and points[i][1] > points[j][1]:
                # If it is, see if it's better to extend the chain ending at j
                dp[i] = max(dp[i], dp[j] + points[i][2])

    # The maximum chain weight is the maximum value in the DP table
    max_chain_weight = max(dp)

    return max_chain_weight

def compare_naive_and_heuristic(input_list):
    n = len(input_list)
    for i in range(1, n + 1):
        subsets = [input_list[:i]]
        for subset in subsets:
            subset_list = list(subset)

            print("Subset Size:", len(subset_list))

            # Run the naive approach
            start_time = time.time()
            original_wis, _ = weighted_increasing_subsequence_iter(subset_list)
            naive_changed = naive_approach(subset_list, original_wis)
            end_time = time.time()
            naive_time = end_time - start_time

            # Run the heuristic algorithm
            start_time = time.time()
            original_wis, ranked_input = weighted_increasing_subsequence_iter(subset_list)
            heuristic_changed, _ = heuristic_algo(subset_list, ranked_input, original_wis)
            end_time = time.time()
            heuristic_time = end_time - start_time

            print("Naive Approach - Number of Items Changed:", naive_changed)
            print("Naive Approach - Execution Time:", naive_time, "seconds")
            print("Heuristic Algorithm - Number of Items Changed:", len(heuristic_changed))
            print("Heuristic Algorithm - Execution Time:", heuristic_time, "seconds")
            print("")


def compare_naive_and_heuristic_graph(input_list):
    n = len(input_list)
    subset_sizes = []
    naive_times = []
    heuristic_times = []
    naive_changed_items = []
    heuristic_changed_items = []

    for i in range(1, n + 1):
        subsets = [input_list[:i]]
        for subset in subsets:
            subset_list = list(subset)
            subset_size = len(subset_list)

            subset_sizes.append(subset_size)

            # Run the naive approach
            start_time = time.time()
            original_wis, _ = maximum_weight_increasing_sequence4(subset_list)
            naive_changed = naive_approach(subset_list, original_wis)
            end_time = time.time()
            naive_time = end_time - start_time

            naive_times.append(naive_time)
            naive_changed_items.append(naive_changed)

            # Run the heuristic algorithm
            start_time = time.time()
            original_wis, ranked_input = maximum_weight_increasing_sequence4(subset_list)
            heuristic_changed, _ = heuristic_algo(subset_list, ranked_input, original_wis)
            end_time = time.time()
            heuristic_time = end_time - start_time

            heuristic_times.append(heuristic_time)
            heuristic_changed_items.append(len(heuristic_changed))

    # Plotting the time comparison graph
    plt.figure()
    plt.plot(subset_sizes, naive_times, label='Naive Approach')
    plt.plot(subset_sizes, heuristic_times, label='Heuristic Algorithm')
    plt.xlabel('Subset Size')
    plt.ylabel('Execution Time (seconds)')
    plt.title('Time Comparison: Naive Approach vs Heuristic Algorithm')
    plt.legend()
    plt.show()

    # Plotting the item change comparison graph
    plt.figure()
    prev_equal_result = False
    prev_naive_heuristic = False
    for i in range(len(subset_sizes)):
        if naive_changed_items[i] == heuristic_changed_items[i]:
            if not prev_equal_result:
                offset = random.uniform(-0.1, 0.1)  # Random offset for perturbation
                plt.plot(subset_sizes[i] + offset, naive_changed_items[i], 'ro', label='Equal Result')
                prev_equal_result = True
            else:
                plt.plot(subset_sizes[i] + offset, naive_changed_items[i], 'ro')
        else:
            if not prev_naive_heuristic:
                plt.plot(subset_sizes[i], naive_changed_items[i], 'o', label='Naive Approach', color='green')
                plt.plot(subset_sizes[i], heuristic_changed_items[i], 'o', label='Heuristic Algorithm', color='blue')
                prev_naive_heuristic = True
            else:
                plt.plot(subset_sizes[i] + offset, naive_changed_items[i], 'ro', color='green')
                plt.plot(subset_sizes[i] + offset, heuristic_changed_items[i], 'ro', color='blue')

    # Set the y-axis to display only natural numbers
    plt.gca().yaxis.set_major_locator(ticker.MaxNLocator(integer=True))

    plt.xlabel('Subset Size')
    plt.ylabel('Number of Items Changed')
    plt.title('Item Change Comparison: Naive Approach vs Heuristic Algorithm')
    plt.legend()
    plt.show()


def check_increasing_elements(input_list, indices):
    for i in range(len(indices) - 1):
        curr_index = indices[i]
        next_index = indices[i + 1]
        curr_element = input_list[curr_index]
        next_element = input_list[next_index]

        if curr_element[0] >= next_element[0] or curr_element[1] >= next_element[1]:
            return False

    return True


def find_haviest(input_list):
    max_w = 1
    for i in range(len(input_list)):
        x, y, w = input_list[i]
        tmp_w = 1
        for j in range(i + 1, len(input_list)):
            x_curr, y_curr, w_curr = input_list[i]
            if x_curr > x and y_curr > y:
                tmp_w += w_curr
            if tmp_w > max_w:
                max_w = tmp_w
    return max_w


def find_heaviest_wieght_naive(input_list):
    input_array = np.array(input_list)
    subsets = list(chain(*map(lambda x: combinations(range(len(input_list)), x), range(0, len(input_list)+1))))
    max_changed = 0
    i = 0
    for subset in subsets:
        flag = check_increasing_elements(input_list, subset)
        if flag and max_changed < len(subset):
            max_changed = len(subset)
    return max_changed


def maximum_weight_increasing_sequence4(points):
    # Sort the points based on the x-coordinate
    points.sort(key=lambda p: p[0])

    # Initialize an array to store the maximum weights
    max_weights = [p[2] for p in points]

    # Initialize an array to store the previous indices
    previous_indices = [-1] * len(points)

    # Traverse the points and update the max_weights and previous_indices arrays
    for i in range(1, len(points)):
        for j in range(i):
            if points[i][0] > points[j][0] and points[i][1] > points[j][1]:
                if max_weights[i] < max_weights[j] + points[i][2]:
                    max_weights[i] = max_weights[j] + points[i][2]
                    previous_indices[i] = j

    # Find the maximum weight and its corresponding sequence for the entire input
    max_weight = max(max_weights)
    max_index = max_weights.index(max_weight)
    sequence = []
    while max_index != -1:
        sequence.append(points[max_index])
        max_index = previous_indices[max_index]
    sequence.reverse()

    # Create a dictionary to store the total weight in the MWIS from each point to the end
    mwis_dict = {}

    # Calculate the total weight in the MWIS from each point to the end
    for i, point in enumerate(points):
        total_weight = sum(p[2] for p in sequence if p[0] >= point[0] and p[1] >= point[1] and p != point)
        mwis_dict[point] = total_weight + point[2]

    return max_weight, mwis_dict


def create_coordinate_list(n):
    coordinate_list = []
    for _ in range(n):
        x = random.random()
        y = random.random()
        coordinate_list.append((x, y, 1))
    return coordinate_list


def reduction_to_third_dimension(points):
    return  [(x, x, w) for x, y, w in sorted(points, key=lambda tup: (tup[1], tup[0]))]


if __name__ == '__main__':
    # input_list = [(4, 4, 1), (2, 3, 1), (5, 6, 1), (1, 1, 1), (7, 8, 1), (8, 9, 1), (10, 10, 1), (3, 2, 1), (6, 7, 1)]
    # input_list = [(4, 4, 1), (2, 3, 1), (5, 6, 1), (1, 1, 1), (7, 8, 1), (8, 9, 1), (10, 10, 1), (3, 2, 1), (6, 7, 1),
    #               (9, 5, 1), (0, 0, 1), (5, 3, 1), (2, 6, 1), (4, 9, 1), (6, 3, 1), (1, 7, 1)]

    for i in range(1, 11):
        for shape in ['square', 'rhombus']:
            print('\n\nResults Of Arbiv heuristic: ' + shape  + '_' + str(i) +'000_samples')
            input_list_main = read_xlsx_file(shape  + '_' + str(i) +'000_samples', i * 1000)
            execute(input_list_main, str( shape  + '_' + str(i) +'000_samples'))



